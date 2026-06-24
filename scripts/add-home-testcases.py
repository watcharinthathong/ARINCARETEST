"""
add-home-testcases.py
เพิ่ม TC-HOME-005 ถึง TC-HOME-008 ใน Test Cases sheet
แล้วเขียนผลเทสจาก test-results-home.json
"""

import json, os, sys, copy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out = sys.stdout.buffer
def p(t): out.write((str(t) + "\n").encode("utf-8"))

BASE  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL = os.path.join(BASE, "Arincare_Telepharmacy_TestCases.xlsx")
JSON  = os.path.join(BASE, "test-results-home.json")

# ── Load results ───────────────────────────────────────────────────────────────
with open(JSON, "r", encoding="utf-8") as f:
    home_results = {r["id"]: r for r in json.load(f)}

# ── Styles ─────────────────────────────────────────────────────────────────────
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(bold=True, color="276221", size=10)
FAIL_FONT = Font(bold=True, color="9C0006", size=10)
NORM_FONT = Font(size=9)
WRAP      = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin      = Side(style="thin", color="CCCCCC")
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Test case definitions for TC-HOME-005 to 008 ──────────────────────────────
# (module, scenario, priority, type, precondition, expected, steps, data, req)
NEW_CASES = [
    (
        "TC-HOME-005", "CMS-Home",
        "แสดงข้อมูลเภสัชกรผู้ควบคุมบน Home",
        "Medium", "Functional",
        "เข้าสู่ระบบ Operator ครบ flow (store → branch → supervisor)",
        "Home page แสดงชื่อ/ข้อมูลเภสัชกรผู้ควบคุมที่เลือก",
        "1. Login เป็น Operator\n2. เลือก Store\n3. เลือก Branch\n4. เลือก Supervisor\n5. ดูข้อมูล supervisor บน Home page",
        "operator@medcare.com / Oper@1234",
        "FR-L02B",
    ),
    (
        "TC-HOME-006", "CMS-Home",
        "ปุ่ม Logout ปรากฏใน Sidebar",
        "High", "Functional",
        "เข้าสู่ระบบแล้ว อยู่หน้า Home",
        "ปุ่ม Logout มองเห็นได้ที่ footer ของ sidebar ซ้าย",
        "1. Login ครบ flow จนถึง Home\n2. ดูที่ footer ของ sidebar ซ้าย",
        "-",
        "-",
    ),
    (
        "TC-HOME-007", "CMS-Home",
        "Logout redirect กลับหน้า Login",
        "High", "Functional",
        "เข้าสู่ระบบแล้ว อยู่หน้า Home",
        "หลัง Logout ระบบ redirect ไปหน้า /login",
        "1. เข้าหน้า Home\n2. คลิกปุ่ม Logout\n3. ยืนยัน confirmation dialog\n4. ตรวจ URL ว่า redirect ไป /login",
        "-",
        "-",
    ),
    (
        "TC-HOME-008", "CMS-Home",
        "Home แสดง Patient Queue และสถานะคิว",
        "High", "Functional",
        "เข้าสู่ระบบ Operator ครบ flow",
        "h1 = 'Patient Queue (คิวผู้ป่วย)', แสดง tab: ทั้งหมด / Active / Waiting / Paused / Closed",
        "1. Login ครบ flow จนถึง Home\n2. ดู h1 heading\n3. ดู queue status tabs",
        "-",
        "Ds-fe-002",
    ),
]

# ── Open workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["Test Cases"]

# ── Find insert position: after TC-HOME-004 (row 25) ──────────────────────────
insert_after = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 1).value == "TC-HOME-004":
        insert_after = row
        break

if insert_after is None:
    p("ERROR: TC-HOME-004 not found"); sys.exit(1)

p(f"Inserting 4 rows after row {insert_after} (TC-HOME-004)")

# ── Shift rows down to make space ──────────────────────────────────────────────
num_new = len(NEW_CASES)
ws.insert_rows(insert_after + 1, amount=num_new)

# ── Copy cell format from TC-HOME-004 as template ─────────────────────────────
template_row = insert_after  # TC-HOME-004

def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = copy.copy(src_cell.font)
        dst_cell.fill      = copy.copy(src_cell.fill)
        dst_cell.border    = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)

# ── Write new test cases ───────────────────────────────────────────────────────
for i, (tc_id, module, scenario, priority, tc_type,
        precond, expected, steps, data, req) in enumerate(NEW_CASES):

    new_row = insert_after + 1 + i
    result  = home_results.get(tc_id, {})
    status  = result.get("status", "")
    actual  = result.get("actualResult", "")
    remark  = result.get("remark", "")
    shots   = result.get("screenshots", [])

    # copy style from template row
    for col in range(1, 14):
        copy_cell_style(ws.cell(template_row, col), ws.cell(new_row, col))

    # col 1-10: test case definition
    vals = [tc_id, module, scenario, priority, tc_type,
            precond, expected, steps, data, req]
    for col, val in enumerate(vals, start=1):
        c = ws.cell(new_row, col)
        c.value = val
        c.border = BORDER
        if col in (1, 4, 5):
            c.alignment = CENTER
        else:
            c.alignment = WRAP

    # col 11: Actual Result
    c11 = ws.cell(new_row, 11)
    c11.value     = actual
    c11.font      = NORM_FONT
    c11.alignment = WRAP
    c11.border    = BORDER

    # col 12: Status
    c12 = ws.cell(new_row, 12)
    c12.value     = status
    c12.alignment = CENTER
    c12.border    = BORDER
    if status == "PASS":
        c12.fill = PASS_FILL
        c12.font = PASS_FONT
    elif status == "FAIL":
        c12.fill = FAIL_FILL
        c12.font = FAIL_FONT

    # col 13: Remark + screenshots
    ss_text = ""
    if shots:
        ss_text = "\n\nScreenshots:\n" + "\n".join(f"  • {s}" for s in shots[:5])
        if len(shots) > 5:
            ss_text += f"\n  ... +{len(shots)-5} more"
    c13 = ws.cell(new_row, 13)
    c13.value     = remark + ss_text
    c13.font      = NORM_FONT
    c13.alignment = WRAP
    c13.border    = BORDER

    ws.row_dimensions[new_row].height = 60

    icon = "✅" if status == "PASS" else "❌" if status == "FAIL" else "➕"
    p(f"  {icon} row {new_row} | {tc_id} | {status} — {scenario}")

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(EXCEL)
p(f"\nSaved: {EXCEL}")
p(f"Added {num_new} rows (TC-HOME-005 to TC-HOME-008)")
