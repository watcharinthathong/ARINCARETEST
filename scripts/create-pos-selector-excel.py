"""
create-pos-selector-excel.py
สร้าง Excel สรุป Selector ของระบบ POS สมัครสมาชิกใหม่
รันด้วย: python scripts/create-pos-selector-excel.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Styles ──────────────────────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="1565C0")
LBLUE_FILL  = PatternFill("solid", fgColor="1976D2")
GREEN_FILL  = PatternFill("solid", fgColor="2E7D32")
PURPLE_FILL = PatternFill("solid", fgColor="6A1B9A")
ORANGE_FILL = PatternFill("solid", fgColor="E65100")
TEAL_FILL   = PatternFill("solid", fgColor="00695C")
GRAY_FILL   = PatternFill("solid", fgColor="455A64")
HEADER_FILL = PatternFill("solid", fgColor="263238")

EVEN_FILL   = PatternFill("solid", fgColor="F5F5F5")
ODD_FILL    = PatternFill("solid", fgColor="FFFFFF")
REQ_FILL    = PatternFill("solid", fgColor="FFF9C4")    # yellow — required
NOTE_FILL   = PatternFill("solid", fgColor="E8F5E9")    # light green — devtools needed

WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=10)
BLACK_BOLD  = Font(bold=True, color="000000", size=10)
BLACK_NORM  = Font(color="000000", size=9)
RED_NORM    = Font(color="C62828", size=9, bold=True)

thin = Side(style="thin", color="B0BEC5")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_header(ws, row, cols, fill, font=WHITE_BOLD):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = CENTER
        c.border = BORDER


def add_row(ws, row, values, fill=None, font=None):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill or (EVEN_FILL if row % 2 == 0 else ODD_FILL)
        c.font = font or BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Login Flow Selectors
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Login Flow"

# Title
ws1.merge_cells("A1:E1")
c = ws1["A1"]
c.value = "POS Login Flow — Selector Reference (pos-stg.arincare.com)"
c.fill = HEADER_FILL
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = CENTER

# Sub-header
ws1.merge_cells("A2:E2")
c = ws1["A2"]
c.value = "ขั้นตอน Login ตั้งแต่หน้า Login จนถึงหน้า Main POS (URL: https://pos-stg.arincare.com/login)"
c.fill = PatternFill("solid", fgColor="37474F")
c.font = Font(color="ECEFF1", size=9)
c.alignment = CENTER

# Headers
ws1.row_dimensions[3].height = 28
ws1.append([])  # row 3 placeholder
ws1.cell(row=3, column=1).value = "หน้าจอ / ขั้นตอน"
ws1.cell(row=3, column=2).value = "ชื่อ Element"
ws1.cell(row=3, column=3).value = "Selector (Playwright)"
ws1.cell(row=3, column=4).value = "ประเภท Element"
ws1.cell(row=3, column=5).value = "Remark"
style_header(ws1, 3, 5, HEADER_FILL)

login_data = [
    # Screen, Element, Selector, Type, Remark
    ("Login Page\n(pos-stg.arincare.com/login)",
     "ช่องกรอก Username / Email",
     'input[name="email"]',
     "input[type=text]",
     'placeholder="อีเมล"'),

    ("Login Page\n(pos-stg.arincare.com/login)",
     "ช่องกรอก Password",
     'input[type="password"]',
     "input[type=password]",
     'placeholder="รหัสผ่าน"'),

    ("Login Page\n(pos-stg.arincare.com/login)",
     "ปุ่มเข้าสู่ระบบ (Login)",
     'button:has-text("เข้าสู่ระบบ")',
     "button[type=submit]",
     "ปุ่มแรก — submit email/password"),

    ("Login Page\n(pos-stg.arincare.com/login)",
     "ปุ่มลืมรหัสผ่าน",
     'button:has-text("ลืมรหัสผ่าน")',
     "button[type=button]",
     ""),

    ("Company/Branch Setup\n(Modal: ตั้งค่า ARINCARE POS)",
     "Dropdown เลือกบริษัท",
     'select[name="companyId"]',
     "select",
     "Native <select> — ใช้ selectOption({label: '....'})"),

    ("Company/Branch Setup\n(Modal: ตั้งค่า ARINCARE POS)",
     "Dropdown เลือกสาขา",
     'select[name="branchId"]',
     "select",
     "Native <select> — ใช้ selectOption({label: '....'})"),

    ("Company/Branch Setup\n(Modal: ตั้งค่า ARINCARE POS)",
     "ปุ่มบันทึกการเปลี่ยนแปลง",
     'button:has-text("บันทึกการเปลี่ยนแปลง")',
     "button",
     "ปุ่มสีแดง — submit company/branch"),

    ("Setup Complete\n(Modal: การตั้งค่า ARINCARE POS เสร็จสิ้น)",
     "ปุ่มเสร็จสิ้น",
     'button:has-text("เสร็จสิ้น")',
     "button",
     "ปุ่มสีเขียว — confirm setup complete"),

    ("Employee Login\n(Overlay ด้านขวา)",
     "ช่องรหัสประจำตัวพนักงาน",
     'input[name="username"]',
     "input[type=text]",
     'placeholder="รหัสประจำตัวพนักงาน"'),

    ("Employee Login\n(Overlay ด้านขวา)",
     "ช่องรหัสผ่านพนักงาน",
     'input[type="password"]',
     "input[type=password]",
     'placeholder="รหัสผ่าน" — ระวัง: ซ้ำกับ Login Page'),

    ("Employee Login\n(Overlay ด้านขวา)",
     "ปุ่มเข้าสู่ระบบ (พนักงาน)",
     'button:has-text("เข้าสู่ระบบ")',
     "button[type=submit]",
     "ปุ่มที่ 2 ในหน้า — ควร scope ใน overlay"),

    ("Popup / Modal (โฆษณา / แจ้งเตือน)",
     "ปุ่มปิด Popup",
     'button:has-text("ปิด")',
     "button",
     "อาจมีหลาย popup — ให้ loop ปิดจนหมด"),

    ("Main POS Page\n(pos-stg.arincare.com/)",
     "ช่องค้นหาลูกค้าสมาชิก",
     '[placeholder="ค้นหาลูกค้าสมาชิก (ctrl + M)"]',
     "input[type=text]",
     "อยู่ด้านขวาของหน้า POS"),

    ("Main POS Page\n(pos-stg.arincare.com/)",
     "ปุ่มสมัครสมาชิกใหม่",
     'button:has-text("สมัครสมาชิกใหม่")',
     "button",
     "เปิด Modal ฟอร์มสมัครสมาชิก"),
]

# Screen section colors
screen_colors = {
    "Login Page": PatternFill("solid", fgColor="E3F2FD"),
    "Company/Branch": PatternFill("solid", fgColor="E8F5E9"),
    "Setup Complete": PatternFill("solid", fgColor="F3E5F5"),
    "Employee Login": PatternFill("solid", fgColor="FFF3E0"),
    "Popup": PatternFill("solid", fgColor="FCE4EC"),
    "Main POS": PatternFill("solid", fgColor="E0F7FA"),
}

row = 4
for screen, element, selector, etype, remark in login_data:
    fill = ODD_FILL
    for key, f in screen_colors.items():
        if key.lower() in screen.lower():
            fill = f
            break
    add_row(ws1, row, [screen, element, selector, etype, remark], fill=fill)
    row += 1

# Column widths
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["C"].width = 48
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 48


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Registration Form — ข้อมูลทั่วไป
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Tab ข้อมูลทั่วไป")

ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "ฟอร์มสมัครสมาชิกใหม่ — Tab: ข้อมูลทั่วไป"
c.fill = BLUE_FILL
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = CENTER

ws2.row_dimensions[2].height = 28
headers = ["ชื่อ Element", "Selector (Playwright)", "ประเภท Element", "Required", "Placeholder / Hint", "Remark"]
for col, h in enumerate(headers, start=1):
    c = ws2.cell(row=2, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = BORDER

tab1_data = [
    ("ชื่อ (First Name)", 'input[name="first_name"]', "input[type=text]", "YES", "เช่น สมชาย", ""),
    ("นามสกุล (Last Name)", 'input[name="last_name"]', "input[type=text]", "YES", "เช่น ใจดี", ""),
    ("เบอร์โทรศัพท์มือถือ", 'input[name="mobile_number"]', "input[type=text]", "YES", "08X-XXX-XXXX", ""),
    ("อีเมล", 'input[name="email"]', "input[type=email]", "NO", "example@email.com", ""),
    ("วันเกิด", '[placeholder="วว/ดด/ปปปป"]', "input[type=text]", "YES", "วว/ดด/ปปปป", "ต้องตรวจสอบด้วย DevTools — ไม่มี name/id"),
    ("เพศ", 'select[name="sex"]', "select", "YES", "เลือกเพศ", ""),
    ("สัญชาติ", 'select[name="nationality"]', "select", "NO", "ไทย (default)", ""),
    ("บัตรประชาชน / Passport", 'input[name="citizen_id"]', "input[type=text]", "NO", "X-XXXX-XXXXX-XX-X", ""),
    ("อาชีพ", 'input[name="occupation"]', "input[type=text]", "NO", "เช่น พนักงานบริษัท", ""),
    ("หมู่เลือด", 'select[name="blood_type"]', "select", "NO", "เลือกหมู่เลือด", ""),
    ("ระดับราคา", 'select[name="price_level"]', "select", "NO", "ตามการตั้งค่าของบริษัท", ""),
    ("ปุ่มยกเลิก", 'button:has-text("ยกเลิก")', "button", "-", "", "ปุ่มสีแดง"),
    ("ปุ่มบันทึก", 'button:has-text("บันทึก")', "button", "-", "", "ปุ่มสีเขียว"),
    ("ปุ่มปิด Modal (×)", '[aria-label="Close"]', "button", "-", "", "ปุ่ม × มุมขวาบน"),
]

row = 3
for item in tab1_data:
    name, sel, etype, req, ph, remark = item
    fill = REQ_FILL if req == "YES" else (ODD_FILL if row % 2 == 0 else EVEN_FILL)
    if remark.startswith("ต้องตรวจสอบ"):
        fill = NOTE_FILL
    for col, val in enumerate([name, sel, etype, req, ph, remark], start=1):
        c = ws2.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = RED_NORM if req == "YES" and col == 4 else BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER
    row += 1

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 48
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 28
ws2.column_dimensions["F"].width = 42


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Registration Form — ข้อมูลใบกำกับภาษี
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Tab ข้อมูลใบกำกับภาษี")

ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "ฟอร์มสมัครสมาชิกใหม่ — Tab: ข้อมูลใบกำกับภาษี"
c.fill = GREEN_FILL
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = CENTER

ws3.row_dimensions[2].height = 28
for col, h in enumerate(headers, start=1):
    c = ws3.cell(row=2, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = BORDER

tab2_data = [
    ("Tab Selector", 'li:has-text("ข้อมูลใบกำกับภาษี")', "tab/li", "-", "", "คลิกเพื่อเปิด Tab 2"),
    ("ชื่อบริษัท", 'input[name="company_name"]', "input[type=text]", "NO", "เช่น บริษัท ABC จำกัด", ""),
    ("เลขประจำตัวผู้เสียภาษี", 'input[name="tax_id"]', "input[type=text]", "NO", "", ""),
    ("ผู้ติดต่อ", 'input[name="contact_name"]', "input[type=text]", "NO", "ชื่อผู้ติดต่อประสานงาน", ""),
    ("เบอร์ติดต่อ", 'input[name="phone_number"]', "input[type=text]", "NO", "", ""),
    ("ชื่อสถานที่", 'input[name="location_name"]', "input[type=text]", "NO", "เช่น สำนักงานใหญ่", ""),
    ("ที่อยู่ 1", 'textarea[name="address1"]', "textarea", "NO", "บ้านเลขที่ ซอย ถนน", ""),
    ("ที่อยู่ 2", 'textarea[name="address2"]', "textarea", "NO", "รายละเอียดเพิ่มเติม (ถ้ามี)", ""),
    ("จังหวัด", 'select[name="province_id"]', "select", "NO", "", "Cascading dropdown — โหลด อำเภอ/เขต หลังเลือก"),
    ("อำเภอ/เขต", 'select[name="city_id"]', "select", "NO", "", "Cascading dropdown — โหลดหลัง จังหวัด"),
    ("ตำบล/แขวง", 'select[name="district_id"]', "select", "NO", "", "Cascading dropdown — โหลดหลัง อำเภอ/เขต"),
    ("รหัสไปรษณีย์", 'input[name="zipcode"]', "input[type=text]", "NO", "", ""),
    ("ปุ่มยกเลิก", 'button:has-text("ยกเลิก")', "button", "-", "", "ปุ่มสีแดง"),
    ("ปุ่มบันทึก", 'button:has-text("บันทึก")', "button", "-", "", "ปุ่มสีเขียว"),
    ("ปุ่มปิด Modal (×)", '[aria-label="Close"]', "button", "-", "", "ปุ่ม × มุมขวาบน"),
]

row = 3
for item in tab2_data:
    name, sel, etype, req, ph, remark = item
    fill = REQ_FILL if req == "YES" else (ODD_FILL if row % 2 == 0 else EVEN_FILL)
    if "Cascading" in remark or remark.startswith("ต้อง"):
        fill = NOTE_FILL
    for col, val in enumerate([name, sel, etype, req, ph, remark], start=1):
        c = ws3.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER
    row += 1

for col in ["A","B","C","D","E","F"]:
    ws3.column_dimensions[col].width = ws2.column_dimensions[col].width


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: Registration Form — หมายเหตุและการแพ้ยา
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Tab หมายเหตุและการแพ้ยา")

ws4.merge_cells("A1:F1")
c = ws4["A1"]
c.value = "ฟอร์มสมัครสมาชิกใหม่ — Tab: หมายเหตุและการแพ้ยา"
c.fill = PURPLE_FILL
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = CENTER

ws4.row_dimensions[2].height = 28
for col, h in enumerate(headers, start=1):
    c = ws4.cell(row=2, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = BORDER

tab3_data = [
    ("Tab Selector", 'li:has-text("หมายเหตุและการแพ้ยา")', "tab/li", "-", "", "คลิกเพื่อเปิด Tab 3"),
    ("ช่องหมายเหตุ (บันทึกข้อมูลเพิ่มเติม)", 'textarea[name="note"]', "textarea", "NO", "บันทึกข้อมูลเพิ่มเติม...", "Textarea เดียวกันกับ Tab 1"),
    ("ปุ่มเพิ่มสินค้าที่มีอาการแพ้ยา", "ต้องตรวจสอบด้วย DevTools เพิ่มเติม", "button", "NO", "", "Tab 3 ถูก popup บัง — ยังไม่สามารถยืนยัน selector ได้"),
    ("ปุ่มเพิ่มชื่อสามัญ (สารที่แพ้)", "ต้องตรวจสอบด้วย DevTools เพิ่มเติม", "button", "NO", "", "Tab 3 ถูก popup บัง — ยังไม่สามารถยืนยัน selector ได้"),
    ("ปุ่มเพิ่มโรคประจำตัว", "ต้องตรวจสอบด้วย DevTools เพิ่มเติม", "button", "NO", "", "Tab 3 ถูก popup บัง — ยังไม่สามารถยืนยัน selector ได้"),
    ("ปุ่มยกเลิก", 'button:has-text("ยกเลิก")', "button", "-", "", "ปุ่มสีแดง"),
    ("ปุ่มบันทึก", 'button:has-text("บันทึก")', "button", "-", "", "ปุ่มสีเขียว"),
    ("ปุ่มปิด Modal (×)", '[aria-label="Close"]', "button", "-", "", "ปุ่ม × มุมขวาบน"),
]

row = 3
for item in tab3_data:
    name, sel, etype, req, ph, remark = item
    fill = NOTE_FILL if "DevTools" in sel or "DevTools" in remark else (ODD_FILL if row % 2 == 0 else EVEN_FILL)
    for col, val in enumerate([name, sel, etype, req, ph, remark], start=1):
        c = ws4.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER
    row += 1

for col in ["A","B","C","D","E","F"]:
    ws4.column_dimensions[col].width = ws2.column_dimensions[col].width


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Summary — ทุก Selector รวมกัน
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary")

ws5.merge_cells("A1:G1")
c = ws5["A1"]
c.value = "Selector Summary — POS สมัครสมาชิกใหม่ (pos-stg.arincare.com)"
c.fill = HEADER_FILL
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = CENTER

ws5.row_dimensions[2].height = 28
sum_headers = ["#", "หน้าจอ", "ชื่อ Element", "Selector (Playwright)", "ประเภท Element", "Required", "Remark"]
for col, h in enumerate(sum_headers, start=1):
    c = ws5.cell(row=2, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = BORDER

all_rows = []

# Login flow
screen_fills = {
    "Login Page": PatternFill("solid", fgColor="E3F2FD"),
    "Company/Branch": PatternFill("solid", fgColor="E8F5E9"),
    "Setup Complete": PatternFill("solid", fgColor="F3E5F5"),
    "Employee Login": PatternFill("solid", fgColor="FFF3E0"),
    "Popup": PatternFill("solid", fgColor="FCE4EC"),
    "Main POS": PatternFill("solid", fgColor="E0F7FA"),
    "Tab: ข้อมูลทั่วไป": PatternFill("solid", fgColor="E3F2FD"),
    "Tab: ข้อมูลใบกำกับภาษี": PatternFill("solid", fgColor="E8F5E9"),
    "Tab: หมายเหตุ": PatternFill("solid", fgColor="F3E5F5"),
}

for screen, element, selector, etype, remark in login_data:
    all_rows.append((screen.split("\n")[0], element, selector, etype, "-", remark))

for element, sel, etype, req, ph, remark in tab1_data:
    all_rows.append(("Tab: ข้อมูลทั่วไป", element, sel, etype, req, remark or ph))

for element, sel, etype, req, ph, remark in tab2_data:
    all_rows.append(("Tab: ข้อมูลใบกำกับภาษี", element, sel, etype, req, remark or ph))

for element, sel, etype, req, ph, remark in tab3_data:
    all_rows.append(("Tab: หมายเหตุและการแพ้ยา", element, sel, etype, req, remark or ph))

row = 3
for i, (screen, element, sel, etype, req, remark) in enumerate(all_rows, start=1):
    fill = ODD_FILL
    for key, f in screen_fills.items():
        if key in screen:
            fill = f
            break
    if "DevTools" in sel or "DevTools" in remark:
        fill = NOTE_FILL
    if req == "YES":
        fill = REQ_FILL
    for col, val in enumerate([i, screen, element, sel, etype, req, remark], start=1):
        c = ws5.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER
    row += 1

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 28
ws5.column_dimensions["C"].width = 32
ws5.column_dimensions["D"].width = 48
ws5.column_dimensions["E"].width = 22
ws5.column_dimensions["F"].width = 12
ws5.column_dimensions["G"].width = 42

# ── Legend ──────────────────────────────────────────────────────────────────
row += 2
ws5.cell(row=row, column=1).value = "Legend:"
ws5.cell(row=row, column=1).font = Font(bold=True, size=10)
row += 1
legend = [
    (REQ_FILL, "สีเหลือง = Required field"),
    (NOTE_FILL, "สีเขียวอ่อน = ต้องตรวจสอบด้วย DevTools เพิ่มเติม"),
    (PatternFill("solid", fgColor="E3F2FD"), "สีฟ้าอ่อน = Login/Tab ข้อมูลทั่วไป"),
    (PatternFill("solid", fgColor="E8F5E9"), "สีเขียวอ่อน = Company setup/Tab ใบกำกับภาษี"),
    (PatternFill("solid", fgColor="F3E5F5"), "สีม่วงอ่อน = Tab หมายเหตุ"),
]
for fill, text in legend:
    ws5.cell(row=row, column=1).fill = fill
    ws5.cell(row=row, column=1).border = BORDER
    ws5.cell(row=row, column=2).value = text
    ws5.cell(row=row, column=2).font = BLACK_NORM
    row += 1

# ── Save ─────────────────────────────────────────────────────────────────────
output = "pos-selector-discovery/POS_Member_Selectors.xlsx"
wb.save(output)
print(f"✅ บันทึก Excel: {output}")
print(f"   - Sheet 'Login Flow'       : {len(login_data)} rows")
print(f"   - Sheet 'Tab ข้อมูลทั่วไป'    : {len(tab1_data)} rows")
print(f"   - Sheet 'Tab ข้อมูลใบกำกับภาษี': {len(tab2_data)} rows")
print(f"   - Sheet 'Tab หมายเหตุและการแพ้ยา': {len(tab3_data)} rows")
print(f"   - Sheet 'Summary'           : {len(all_rows)} rows total")
