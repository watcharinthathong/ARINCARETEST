"""
create-telepharmacy-selector-excel.py
สร้าง Excel สรุป Selector ของระบบ Telepharmacy CMS
ครอบคลุม: Login → Select Store → Select Branch → Select Supervisor → Home

รันด้วย: python scripts/create-telepharmacy-selector-excel.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()

# ─── Styles ──────────────────────────────────────────────────────────────────
DARK_FILL    = PatternFill("solid", fgColor="1A237E")   # Indigo dark - title
BLUE_FILL    = PatternFill("solid", fgColor="1565C0")
GREEN_FILL   = PatternFill("solid", fgColor="2E7D32")
TEAL_FILL    = PatternFill("solid", fgColor="00695C")
PURPLE_FILL  = PatternFill("solid", fgColor="6A1B9A")
BROWN_FILL   = PatternFill("solid", fgColor="4E342E")
HEADER_FILL  = PatternFill("solid", fgColor="263238")

# Row highlight by page
LOGIN_FILL   = PatternFill("solid", fgColor="E3F2FD")   # light blue - Login
STORE_FILL   = PatternFill("solid", fgColor="E8F5E9")   # light green - Store
BRANCH_FILL  = PatternFill("solid", fgColor="FFF3E0")   # light orange - Branch
SUPER_FILL   = PatternFill("solid", fgColor="F3E5F5")   # light purple - Supervisor
HOME_FILL    = PatternFill("solid", fgColor="E0F7FA")   # light teal - Home
WARN_FILL    = PatternFill("solid", fgColor="E8F5E9")   # light green - needs verify

EVEN_FILL    = PatternFill("solid", fgColor="F5F5F5")
ODD_FILL     = PatternFill("solid", fgColor="FFFFFF")

WHITE_BOLD   = Font(bold=True, color="FFFFFF", size=10)
BLACK_BOLD   = Font(bold=True, color="000000", size=10)
BLACK_NORM   = Font(color="000000", size=9)
NOTE_FONT    = Font(color="BF360C", size=9, italic=True)   # orange-red for unverified

thin = Side(style="thin", color="B0BEC5")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_title(ws, row, cols, fill, text, size=13):
    ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = Font(bold=True, color="FFFFFF", size=size)
    c.alignment = CENTER


def style_header(ws, row, headers, fill=HEADER_FILL):
    ws.row_dimensions[row].height = 26
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = fill
        c.font = WHITE_BOLD
        c.alignment = CENTER
        c.border = BORDER


def add_row(ws, row, values, fill=ODD_FILL, is_note=False):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = NOTE_FONT if (is_note and col in (3,)) else BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER


def add_section_label(ws, row, cols, text, fill):
    """Bold section separator row spanning all columns"""
    ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = LEFT
    c.border = BORDER
    ws.row_dimensions[row].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Login Flow — ทุก Selector ตาม Flow
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Login Flow"

style_title(ws1, 1, 6, DARK_FILL,
    "Telepharmacy CMS — Selector Reference  (telepharmacy-cms.vercel.app)", 13)

ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value = "Login Flow: /login → /select-store → /select-branch → /select-supervisor → /home"
c.fill = PatternFill("solid", fgColor="37474F")
c.font = Font(color="ECEFF1", size=9)
c.alignment = CENTER

HEADERS_FLOW = ["หน้าจอ / ขั้นตอน", "ชื่อ Element", "Selector (Playwright)",
                "ประเภท", "สถานะ Selector", "Remark"]
style_header(ws1, 3, HEADERS_FLOW)

# (screen, element, selector, type, verified, remark)
FLOW_DATA = [
    # ── Login Page (/login) ────────────────────────────────────────────────────
    ("__SECTION__", "LOGIN PAGE  (/login)", "", "", "", ""),

    ("Login Page\n(/login)",
     "ช่องกรอก Username / Email",
     'input[type="text"]',
     "input",
     "✅ ยืนยันแล้ว",
     'placeholder: "Enter your username" — ไม่มี name/id'),

    ("Login Page\n(/login)",
     "ช่องกรอก Password",
     'input[type="password"]',
     "input",
     "✅ ยืนยันแล้ว",
     'placeholder: "Enter your password"'),

    ("Login Page\n(/login)",
     "ปุ่ม Sign In (Submit)",
     'button[type="submit"]',
     "button",
     "✅ ยืนยันแล้ว",
     "ปุ่มหลัก — disabled เมื่อช่องว่าง"),

    ("Login Page\n(/login)",
     "ปุ่ม Show/Hide Password (Eye icon)",
     'button[type="button"]',
     "button",
     "✅ ยืนยันแล้ว",
     "ปุ่มแรกในหน้า — click เพื่อ toggle type=text/password\nหลัง Show: input.nth(1) type=text\nหลัง Hide: input.nth(1) type=password"),

    ("Login Page\n(/login)",
     "Error Message (login ล้มเหลว)",
     'text=Invalid username or password.',
     "text node",
     "✅ ยืนยันแล้ว",
     'body ประกอบด้วย "Invalid username or password." — ตรวจด้วย body.innerText()'),

    # ── Select Store (/select-store) ──────────────────────────────────────────
    ("__SECTION__", "SELECT STORE  (/select-store)", "", "", "", ""),

    ("Select Store\n(/select-store)",
     "หัวข้อหน้า",
     'text=เลือกร้านค้า',
     "text node",
     "✅ ยืนยันแล้ว",
     "ตรวจด้วย page.url().includes('select-store')"),

    ("Select Store\n(/select-store)",
     "การ์ดร้านค้า (Watcharin TestTest)",
     'text=Watcharin TestTest',
     "div/card",
     "✅ ยืนยันแล้ว",
     ".first() — คลิกเพื่อ highlight/select\nหลังคลิก: button ถัดไป จะ enable"),

    ("Select Store\n(/select-store)",
     "การ์ดร้านค้า (generic — ร้านใดก็ได้)",
     '[class*="card"]:not(:has-text("ถัดไป")):not(:has-text("ย้อนกลับ"))',
     "div/card",
     "⚠️ ตรวจสอบ DevTools",
     "Selector อนุมาน — ยืนยันด้วย DevTools ก่อนใช้ใน production test"),

    ("Select Store\n(/select-store)",
     "ปุ่มถัดไป (หลังเลือกร้าน)",
     'button:has-text("ถัดไป"):not([disabled])',
     "button",
     "✅ ยืนยันแล้ว",
     "ต้องเลือกร้านก่อน ถึงจะ enable\nใช้ :not([disabled]) เพื่อหลีกเลี่ยง disabled state"),

    # ── Select Branch (/select-branch) ────────────────────────────────────────
    ("__SECTION__", "SELECT BRANCH  (/select-branch)", "", "", "", ""),

    ("Select Branch\n(/select-branch)",
     "หัวข้อหน้า",
     'text=เลือกสาขาที่ทำงาน',
     "text node",
     "✅ ยืนยันแล้ว",
     "ตรวจด้วย page.url().includes('select-branch')"),

    ("Select Branch\n(/select-branch)",
     "ชื่อร้านที่เลือก (breadcrumb)",
     'text=Watcharin TestTest',
     "text node",
     "✅ ยืนยันแล้ว",
     "แสดงชื่อร้านพร้อม '· กรุณาเลือกสาขาที่คุณประจำการอยู่'"),

    ("Select Branch\n(/select-branch)",
     "การ์ดสาขา สำนักงานใหญ่ (BTCH00001)",
     'text=สำนักงานใหญ่',
     "div/card",
     "✅ ยืนยันแล้ว",
     ".first() — รหัสสาขา BTCH00001"),

    ("Select Branch\n(/select-branch)",
     "การ์ดสาขา Vteg company (BTCH00002)",
     'text=Vteg company',
     "div/card",
     "✅ ยืนยันแล้ว",
     "รหัสสาขา BTCH00002"),

    ("Select Branch\n(/select-branch)",
     "ปุ่มถัดไป (หลังเลือกสาขา)",
     'button:has-text("ถัดไป"):not([disabled])',
     "button",
     "✅ ยืนยันแล้ว",
     "⚠️ DISABLED จนกว่าจะเลือกสาขา — ต้องคลิกการ์ดสาขาก่อน\nClass ขณะ disabled: disabled:cursor-not-allowed disabled:opacity-50"),

    ("Select Branch\n(/select-branch)",
     "ปุ่มย้อนกลับ",
     'button:has-text("ย้อนกลับ")',
     "button",
     "✅ ยืนยันแล้ว",
     "กลับไปหน้า select-store"),

    # ── Select Supervising Pharmacist (/select-supervisor) ────────────────────
    ("__SECTION__", "SELECT SUPERVISING PHARMACIST  (/select-supervisor)", "", "", "", ""),

    ("Select Supervisor\n(/select-supervisor)",
     "หัวข้อหน้า",
     'text=เลือกเภสัชกรผู้ควบคุม',
     "text node",
     "✅ ยืนยันแล้ว",
     "ตรวจด้วย page.url().includes('select-supervisor')\nหรือ body.includes('เภสัชกรผู้ควบคุม')"),

    ("Select Supervisor\n(/select-supervisor)",
     "การ์ดเภสัชกรผู้ควบคุม",
     '[class*="card"]:not(:has-text("ถัดไป")):not(:has-text("ย้อนกลับ"))',
     "div/card",
     "⚠️ ตรวจสอบ DevTools",
     "Selector อนุมาน — ยืนยันด้วย DevTools\nทดสอบด้วย .first().click() แล้วดูว่า button ถัดไป enable"),

    ("Select Supervisor\n(/select-supervisor)",
     "ชื่อเภสัชกร (text-based)",
     'text=[ชื่อเภสัชกร]',
     "text node",
     "⚠️ ตรวจสอบ DevTools",
     "ระบุชื่อเภสัชกรที่ต้องการเลือก เช่น\npage.locator('text=นพ. สมชาย').click()"),

    ("Select Supervisor\n(/select-supervisor)",
     "ปุ่มยืนยัน / ถัดไป",
     'button:has-text("ถัดไป"):not([disabled])',
     "button",
     "⚠️ ตรวจสอบ DevTools",
     "คาดว่า pattern เดียวกับหน้า store/branch\nต้องเลือก supervisor ก่อนถึง enable"),

    ("Select Supervisor\n(/select-supervisor)",
     "ปุ่มย้อนกลับ",
     'button:has-text("ย้อนกลับ")',
     "button",
     "⚠️ ตรวจสอบ DevTools",
     "กลับไปหน้า select-branch"),

    # ── Home Page (/home หรือ /) ───────────────────────────────────────────────
    ("__SECTION__", "HOME PAGE  (หลัง login สำเร็จ)", "", "", "", ""),

    ("Home Page\n(/home หรือ /)",
     "URL หลัง login สำเร็จ",
     "page.url()",
     "URL check",
     "⚠️ ตรวจสอบ DevTools",
     "ต้องยืนยัน URL ที่ redirect หลัง select-supervisor\nตรวจด้วย !page.url().includes('/select-')"),

    ("Home Page\n(/home หรือ /)",
     "หัวข้อหน้า / Dashboard heading",
     "ต้องตรวจสอบด้วย DevTools",
     "text/heading",
     "⚠️ ตรวจสอบ DevTools",
     "ยังไม่สามารถเข้าถึงได้ (pharmacist login ไม่สำเร็จ)\noperator flow จบที่ select-supervisor"),

    ("Home Page\n(/home หรือ /)",
     "Navigation Menu / Sidebar",
     "ต้องตรวจสอบด้วย DevTools",
     "nav/ul/div",
     "⚠️ ตรวจสอบ DevTools",
     "ตรวจสอบด้วย DevTools หลัง pharmacist login สำเร็จ"),

    ("Home Page\n(/home หรือ /)",
     "ชื่อผู้ใช้ / Profile area",
     "ต้องตรวจสอบด้วย DevTools",
     "div/span",
     "⚠️ ตรวจสอบ DevTools",
     "แสดงชื่อ role ของผู้ใช้ที่ login"),
]

PAGE_FILLS = {
    "Login Page":       LOGIN_FILL,
    "Select Store":     STORE_FILL,
    "Select Branch":    BRANCH_FILL,
    "Select Supervisor": SUPER_FILL,
    "Home Page":        HOME_FILL,
}

SECTION_FILLS = {
    "LOGIN PAGE":      PatternFill("solid", fgColor="1565C0"),
    "SELECT STORE":    PatternFill("solid", fgColor="2E7D32"),
    "SELECT BRANCH":   PatternFill("solid", fgColor="E65100"),
    "SELECT SUPERVIS": PatternFill("solid", fgColor="6A1B9A"),
    "HOME PAGE":       PatternFill("solid", fgColor="00695C"),
}

row = 4
for item in FLOW_DATA:
    screen, element, selector, etype, status, remark = item
    if screen == "__SECTION__":
        sec_fill = ODD_FILL
        for key, f in SECTION_FILLS.items():
            if element.startswith(key):
                sec_fill = f
                break
        add_section_label(ws1, row, 6, f"  ▶  {element}", sec_fill)
        row += 1
        continue

    row_fill = ODD_FILL
    for key, f in PAGE_FILLS.items():
        if key in screen:
            row_fill = f
            break

    is_note = "DevTools" in selector or "DevTools" in remark
    add_row(ws1, row, [screen, element, selector, etype, status, remark],
            fill=row_fill, is_note=is_note)
    row += 1

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 35
ws1.column_dimensions["C"].width = 52
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 20
ws1.column_dimensions["F"].width = 52


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Login Page Detail
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Login Page")

style_title(ws2, 1, 6, BLUE_FILL,
    "Login Page — /login  (telepharmacy-cms.vercel.app/login)", 12)

HEADERS_DETAIL = ["ชื่อ Element", "Selector (Playwright)", "ประเภท", "สถานะ", "หมายเหตุ / ค่าที่คาดหวัง", "Test Case อ้างอิง"]
style_header(ws2, 2, HEADERS_DETAIL)

LOGIN_DETAIL = [
    ("Username Input",
     'input[type="text"]',
     "input[type=text]",
     "✅ ยืนยันแล้ว",
     'placeholder="Enter your username" — ไม่มี name/id attribute',
     "TC-AUTH-001, 002, 003, 004, 005"),

    ("Password Input",
     'input[type="password"]',
     "input[type=password]",
     "✅ ยืนยันแล้ว",
     'placeholder="Enter your password"\nเมื่อกด Show: type เปลี่ยนเป็น text',
     "TC-AUTH-001..006"),

    ("Password (เข้าถึงโดย index)",
     'page.locator("input").nth(1)',
     "input (2nd input)",
     "✅ ยืนยันแล้ว",
     "เป็น input ตัวที่ 2 ในหน้า\nใช้เมื่อ type เปลี่ยนหลัง Show/Hide",
     "TC-AUTH-006"),

    ("Sign In Button",
     'button[type="submit"]',
     "button[type=submit]",
     "✅ ยืนยันแล้ว",
     'text: "Sign In"\ndisabled เมื่อ username/password ว่าง',
     "TC-AUTH-001..005"),

    ("Show/Hide Password Toggle",
     'button[type="button"]',
     "button[type=button]",
     "✅ ยืนยันแล้ว",
     "ปุ่มแรกในหน้า (Eye icon)\nใช้ .first() เพื่อระบุ\nClick 1: show → input.nth(1).getAttribute('type') = 'text'\nClick 2: hide → type = 'password'",
     "TC-AUTH-006"),

    ("Error Message",
     'body',
     "text (innerText)",
     "✅ ยืนยันแล้ว",
     '"Invalid username or password." — ปรากฏใน body.innerText()\nตรวจด้วย /invalid|ไม่ถูกต้อง/i.test(body)',
     "TC-AUTH-003, 004"),

    ("Page copyright text",
     'text=© 2026 Telepharmacy. All Rights Reserved.',
     "text node",
     "✅ ยืนยันแล้ว",
     "ตรวจสอบว่าอยู่หน้า login ถูกต้อง",
     "-"),

    ("Page Title",
     'text=PHARMACIST CONTROL CENTER',
     "heading",
     "✅ ยืนยันแล้ว",
     "Heading ในหน้า Login",
     "-"),
]

row = 3
for item in LOGIN_DETAIL:
    fill = LOGIN_FILL if row % 2 == 0 else PatternFill("solid", fgColor="DDEEFF")
    add_row(ws2, row, list(item), fill=fill)
    row += 1

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 48
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 52
ws2.column_dimensions["F"].width = 30


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Store & Branch Selection
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Select Store + Branch")

style_title(ws3, 1, 6, GREEN_FILL,
    "Select Store & Branch — /select-store  →  /select-branch", 12)

style_header(ws3, 2, HEADERS_DETAIL)

STORE_BRANCH_DATA = [
    # Section: Store
    ("__SECTION__", "SELECT STORE  (/select-store)", "", "", "", ""),

    ("Page URL pattern",
     "page.url().includes('select-store')",
     "URL check",
     "✅ ยืนยันแล้ว",
     "ใช้ตรวจว่าอยู่หน้า select-store ก่อน interact",
     "TC-AUTH-002, 012"),

    ("Store Card — Watcharin TestTest",
     'page.locator("text=Watcharin TestTest").first()',
     "div/card (text match)",
     "✅ ยืนยันแล้ว",
     "คลิกเพื่อ select ร้านนี้\nหลังคลิก: ปุ่มถัดไป จะ enable",
     "TC-AUTH-002, 012"),

    ("Store Card — generic (ร้านใดก็ได้)",
     '[class*="cursor-pointer"]:not(button)',
     "div/card",
     "⚠️ ตรวจสอบ DevTools",
     "Selector อนุมาน — ยังไม่ยืนยัน\nอาจใช้ page.locator('.store-card').first()",
     "-"),

    ("Next Button (ถัดไป) — enabled",
     'button:has-text("ถัดไป"):not([disabled])',
     "button",
     "✅ ยืนยันแล้ว",
     ":not([disabled]) สำคัญมาก — button จะ disabled จนกว่าจะเลือกร้าน\nClass disabled: disabled:cursor-not-allowed disabled:opacity-50",
     "TC-AUTH-002, 012"),

    # Section: Branch
    ("__SECTION__", "SELECT BRANCH  (/select-branch)", "", "", "", ""),

    ("Page URL pattern",
     "page.url().includes('select-branch')",
     "URL check",
     "✅ ยืนยันแล้ว",
     "ใช้ตรวจว่าอยู่หน้า select-branch ก่อน interact",
     "TC-AUTH-002, 012"),

    ("Page Heading",
     'text=เลือกสาขาที่ทำงาน',
     "heading/text",
     "✅ ยืนยันแล้ว",
     "หัวข้อหน้า select-branch",
     "-"),

    ("Store name breadcrumb",
     'text=Watcharin TestTest',
     "text node",
     "✅ ยืนยันแล้ว",
     "แสดงชื่อร้านที่เลือก: 'Watcharin TestTest · กรุณาเลือกสาขาที่คุณประจำการอยู่'",
     "-"),

    ("Branch Card — สำนักงานใหญ่",
     'page.locator("text=สำนักงานใหญ่").first()',
     "div/card (text match)",
     "✅ ยืนยันแล้ว",
     "รหัสสาขา: BTCH00001\nคลิกเพื่อ select → ปุ่มถัดไป enable",
     "TC-AUTH-002, 012"),

    ("Branch Card — Vteg company",
     'page.locator("text=Vteg company").first()',
     "div/card (text match)",
     "✅ ยืนยันแล้ว",
     "รหัสสาขา: BTCH00002",
     "-"),

    ("Next Button (ถัดไป) — enabled",
     'button:has-text("ถัดไป"):not([disabled])',
     "button",
     "✅ ยืนยันแล้ว",
     "⚠️ DISABLED จนกว่าจะเลือกสาขา\nต้องคลิก branch card ก่อนเสมอ",
     "TC-AUTH-002, 012"),

    ("Back Button (ย้อนกลับ)",
     'button:has-text("ย้อนกลับ")',
     "button",
     "✅ ยืนยันแล้ว",
     "กลับไปหน้า select-store",
     "-"),
]

row = 3
for item in STORE_BRANCH_DATA:
    screen, element, selector, etype, remark, tc = item
    if screen == "__SECTION__":
        sec_fill = PatternFill("solid", fgColor="2E7D32") if "STORE" in element else PatternFill("solid", fgColor="E65100")
        add_section_label(ws3, row, 6, f"  ▶  {element}", sec_fill)
        row += 1
        continue
    fill = STORE_FILL if "store" in element.lower() or "Store" in selector or "URL" in etype else BRANCH_FILL
    is_note = "DevTools" in remark or "อนุมาน" in remark
    add_row(ws3, row, [element, selector, etype, screen, remark, tc], fill=fill, is_note=is_note)
    row += 1

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 52
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 52
ws3.column_dimensions["F"].width = 22


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: Select Supervisor & Home
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Supervisor + Home")

style_title(ws4, 1, 6, PURPLE_FILL,
    "Select Supervising Pharmacist (/select-supervisor) + Home Page", 12)

style_header(ws4, 2, HEADERS_DETAIL)

SUPER_HOME_DATA = [
    # Section: Supervisor
    ("__SECTION__", "SELECT SUPERVISING PHARMACIST  (/select-supervisor)", "", "", "", ""),

    ("Page URL pattern",
     "page.url().includes('select-supervisor')",
     "URL check",
     "✅ ยืนยันแล้ว",
     "Operator flow เข้าหน้านี้หลัง select-branch สำเร็จ\n(ยืนยันจาก TC-AUTH-012 PASS)",
     "TC-AUTH-012"),

    ("Page Heading",
     'text=เลือกเภสัชกรผู้ควบคุม',
     "heading/text",
     "✅ ยืนยันแล้ว",
     "ยืนยันจาก body.innerText() ใน TC-AUTH-012",
     "TC-AUTH-012"),

    ("Pharmacist Card (generic)",
     "ต้องตรวจสอบด้วย DevTools",
     "div/card",
     "⚠️ ตรวจสอบ DevTools",
     "แนะนำ: เปิด DevTools > Inspector บนหน้า /select-supervisor\nดู class ของ card ที่ต้องคลิก",
     "-"),

    ("Pharmacist Card (text-based)",
     'text=[ชื่อเภสัชกร]',
     "text match",
     "⚠️ ตรวจสอบ DevTools",
     "ตัวอย่าง: page.locator('text=ภก. สมชาย').first().click()\nต้องทราบชื่อเภสัชกรที่มีในระบบ test",
     "-"),

    ("Confirm/Next Button",
     'button:has-text("ถัดไป"):not([disabled])',
     "button",
     "⚠️ ตรวจสอบ DevTools",
     "คาดว่า pattern เดียวกับหน้าก่อนหน้า\nต้องเลือก pharmacist ก่อนถึง enable",
     "-"),

    ("Back Button",
     'button:has-text("ย้อนกลับ")',
     "button",
     "⚠️ ตรวจสอบ DevTools",
     "กลับไปหน้า select-branch",
     "-"),

    # Section: Home
    ("__SECTION__", "HOME PAGE  (หลัง login flow สำเร็จทั้งหมด)", "", "", "", ""),

    ("URL ที่คาดหวัง",
     "!page.url().includes('/select-')",
     "URL check",
     "⚠️ ตรวจสอบ DevTools",
     "URL final หลัง complete flow ยังไม่ทราบ\nอาจเป็น /home, /dashboard, / หรืออื่นๆ",
     "-"),

    ("Dashboard Heading",
     "ต้องตรวจสอบด้วย DevTools",
     "heading",
     "⚠️ ตรวจสอบ DevTools",
     "เข้าด้วย pharmacist ที่ login สำเร็จ\nดู DOM Inspector ของ heading หลัก",
     "TC-AUTH-011"),

    ("Navigation Menu / Sidebar",
     "ต้องตรวจสอบด้วย DevTools",
     "nav / ul / aside",
     "⚠️ ตรวจสอบ DevTools",
     "Menu items ขึ้นอยู่กับ role (pharmacist vs operator)\nตรวจสอบด้วย page.locator('nav').innerHTML()",
     "TC-AUTH-011"),

    ("User Profile / Role display",
     "ต้องตรวจสอบด้วย DevTools",
     "div / span",
     "⚠️ ตรวจสอบ DevTools",
     "แสดงชื่อผู้ใช้และ role ที่ login\nหาด้วย page.locator('[class*=\"user\"], [class*=\"profile\"]')",
     "-"),

    ("Logout Button",
     'button:has-text("ออกจากระบบ")',
     "button",
     "⚠️ ตรวจสอบ DevTools",
     "Placeholder — ยืนยันข้อความปุ่มด้วย DevTools\nอาจเป็น 'Logout', 'Sign Out', 'ออกจากระบบ'",
     "TC-AUTH-020"),
]

row = 3
for item in SUPER_HOME_DATA:
    screen, element, selector, etype, remark, tc = item
    if screen == "__SECTION__":
        sec_fill = PURPLE_FILL if "SUPER" in element else TEAL_FILL
        add_section_label(ws4, row, 6, f"  ▶  {element}", sec_fill)
        row += 1
        continue
    fill = SUPER_FILL if "supervisor" in element.lower() or "Pharmacist" in element else HOME_FILL
    is_note = "DevTools" in selector or "DevTools" in remark
    add_row(ws4, row, [element, selector, etype, screen, remark, tc], fill=fill, is_note=is_note)
    row += 1

ws4.column_dimensions["A"].width = 32
ws4.column_dimensions["B"].width = 48
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 52
ws4.column_dimensions["F"].width = 22


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Playwright Code Snippets
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Code Snippets")

style_title(ws5, 1, 2, DARK_FILL,
    "Playwright Code Snippets — Telepharmacy CMS Login Flow", 12)

style_header(ws5, 2,
    ["ขั้นตอน", "Code (TypeScript / Playwright)"])

SNIPPETS = [
    ("1. เปิดหน้า Login",
     "await page.goto('https://telepharmacy-cms.vercel.app/login', { waitUntil: 'networkidle' });\nawait page.setViewportSize({ width: 1920, height: 1080 });"),

    ("2. กรอก Username",
     "await page.locator('input[type=\"text\"]').fill('operator@medcare.com');"),

    ("3. กรอก Password",
     "await page.locator('input[type=\"password\"]').fill('Oper@1234');"),

    ("4. กด Sign In",
     "await page.locator('button[type=\"submit\"]').click();\nawait page.waitForTimeout(4000);"),

    ("5. เลือกร้านค้า (Select Store)",
     "if (page.url().includes('select-store')) {\n  await page.locator('text=Watcharin TestTest').first().click();\n  await page.waitForTimeout(600);\n  await page.locator('button:has-text(\"ถัดไป\"):not([disabled])').first().click();\n  await page.waitForTimeout(3000);\n}"),

    ("6. เลือกสาขา (Select Branch)",
     "if (page.url().includes('select-branch')) {\n  await page.locator('text=สำนักงานใหญ่').first().click();  // เลือกก่อน — ปุ่มถัดไป disabled!\n  await page.waitForTimeout(600);\n  await page.locator('button:has-text(\"ถัดไป\"):not([disabled])').first().click();\n  await page.waitForTimeout(3000);\n}"),

    ("7. ตรวจสอบหน้า Supervisor",
     "const hasSupervisor =\n  page.url().includes('select-supervisor') ||\n  (await page.locator('body').innerText()).includes('เภสัชกรผู้ควบคุม');"),

    ("8. Show/Hide Password Toggle",
     "await page.locator('input[type=\"password\"]').fill('Pharm@1234');\nawait page.locator('button[type=\"button\"]').first().click();  // Show\nconst typeAfterShow = await page.locator('input').nth(1).getAttribute('type');  // 'text'\nawait page.locator('button[type=\"button\"]').first().click();  // Hide\nconst typeAfterHide = await page.locator('input').nth(1).getAttribute('type');  // 'password'"),

    ("9. Helper: completeStoreFlow()",
     "async function completeStoreFlow(page: Page) {\n  if (page.url().includes('select-store')) {\n    await page.locator('text=Watcharin TestTest').first().click();\n    await page.waitForTimeout(600);\n    await page.locator('button:has-text(\"ถัดไป\"):not([disabled])').first().click();\n    await page.waitForTimeout(3000);\n  }\n  if (page.url().includes('select-branch')) {\n    await page.locator('text=สำนักงานใหญ่').first().click();\n    await page.waitForTimeout(600);\n    await page.locator('button:has-text(\"ถัดไป\"):not([disabled])').first().click();\n    await page.waitForTimeout(3000);\n  }\n}"),

    ("10. ตรวจ Error Message",
     "const body = await page.locator('body').innerText();\nconst hasError = /invalid|ไม่ถูกต้อง/i.test(body);\nconst errLine = body.split('\\n').find(l => /invalid|ไม่ถูก/i.test(l)) ?? '';"),

    ("11. Screenshot",
     "await page.screenshot({\n  path: 'screenshots/login/TC-AUTH-001_01.png',\n  fullPage: true\n});"),
]

row = 3
for step, code in SNIPPETS:
    ws5.row_dimensions[row].height = max(18, code.count("\n") * 14 + 14)
    c1 = ws5.cell(row=row, column=1, value=step)
    c1.fill = LOGIN_FILL if row % 2 == 0 else EVEN_FILL
    c1.font = BLACK_BOLD
    c1.alignment = LEFT
    c1.border = BORDER

    c2 = ws5.cell(row=row, column=2, value=code)
    c2.fill = PatternFill("solid", fgColor="FAFAFA")
    c2.font = Font(name="Courier New", size=9, color="1A237E")
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border = BORDER
    row += 1

ws5.column_dimensions["A"].width = 32
ws5.column_dimensions["B"].width = 90


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6: Summary — ทุก Selector
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Summary")

style_title(ws6, 1, 7, DARK_FILL,
    "Summary — Telepharmacy CMS Selectors (ทุก Element รวมกัน)", 13)

SUM_HEADERS = ["#", "หน้า", "ชื่อ Element", "Selector (Playwright)", "ประเภท", "สถานะ", "Test Case"]
style_header(ws6, 2, SUM_HEADERS)

SUMMARY_DATA = [
    # Login
    ("Login", "Username Input",           'input[type="text"]',                            "input",     "✅ ยืนยัน", "TC-AUTH-001..005"),
    ("Login", "Password Input",           'input[type="password"]',                        "input",     "✅ ยืนยัน", "TC-AUTH-001..006"),
    ("Login", "Sign In Button",           'button[type="submit"]',                         "button",    "✅ ยืนยัน", "TC-AUTH-001..005"),
    ("Login", "Show/Hide Toggle",         'button[type="button"]',                         "button",    "✅ ยืนยัน", "TC-AUTH-006"),
    ("Login", "Error Message",            'body (innerText contains "Invalid username")',   "text",      "✅ ยืนยัน", "TC-AUTH-003,004"),
    # Select Store
    ("Select Store", "Store Card Text",   'text=Watcharin TestTest',                       "text/card", "✅ ยืนยัน", "TC-AUTH-002,012"),
    ("Select Store", "Next Button",       'button:has-text("ถัดไป"):not([disabled])',      "button",    "✅ ยืนยัน", "TC-AUTH-002,012"),
    # Select Branch
    ("Select Branch", "Branch Card",      'text=สำนักงานใหญ่',                            "text/card", "✅ ยืนยัน", "TC-AUTH-002,012"),
    ("Select Branch", "Branch Card 2",    'text=Vteg company',                             "text/card", "✅ ยืนยัน", "-"),
    ("Select Branch", "Next Button",      'button:has-text("ถัดไป"):not([disabled])',      "button",    "✅ ยืนยัน", "TC-AUTH-002,012"),
    ("Select Branch", "Back Button",      'button:has-text("ย้อนกลับ")',                  "button",    "✅ ยืนยัน", "-"),
    # Select Supervisor
    ("Supervisor",    "Page Heading",     'text=เลือกเภสัชกรผู้ควบคุม',                  "text",      "✅ ยืนยัน", "TC-AUTH-012"),
    ("Supervisor",    "Pharmacist Card",  "ต้องตรวจสอบด้วย DevTools",                     "div/card",  "⚠️ ยังไม่ยืนยัน", "-"),
    ("Supervisor",    "Confirm Button",   'button:has-text("ถัดไป"):not([disabled])',      "button",    "⚠️ ยังไม่ยืนยัน", "-"),
    # Home
    ("Home",          "Page URL",         "!page.url().includes('/select-')",              "URL check", "⚠️ ยังไม่ยืนยัน", "TC-AUTH-011"),
    ("Home",          "Dashboard",        "ต้องตรวจสอบด้วย DevTools",                     "heading",   "⚠️ ยังไม่ยืนยัน", "TC-AUTH-011"),
    ("Home",          "Logout Button",    'button:has-text("ออกจากระบบ")',                 "button",    "⚠️ ยังไม่ยืนยัน", "TC-AUTH-020"),
]

PAGE_ROW_FILLS = {
    "Login":        LOGIN_FILL,
    "Select Store": STORE_FILL,
    "Select Branch": BRANCH_FILL,
    "Supervisor":   SUPER_FILL,
    "Home":         HOME_FILL,
}

row = 3
for i, (page, element, selector, etype, status, tc) in enumerate(SUMMARY_DATA, start=1):
    fill = PAGE_ROW_FILLS.get(page, ODD_FILL)
    is_note = "DevTools" in selector
    values = [i, page, element, selector, etype, status, tc]
    for col, val in enumerate(values, start=1):
        c = ws6.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = NOTE_FONT if (is_note and col == 4) else (
                 Font(color="1A237E", size=9, bold=True) if col == 6 and "✅" in str(val) else
                 Font(color="BF360C", size=9, bold=True) if col == 6 and "⚠️" in str(val) else
                 BLACK_NORM)
        c.alignment = LEFT
        c.border = BORDER
    row += 1

# Legend
row += 1
ws6.cell(row=row, column=1).value = "Legend:"
ws6.cell(row=row, column=1).font = Font(bold=True, size=10)
row += 1
for fill, label in [
    (LOGIN_FILL,   "🟦 Login Page"),
    (STORE_FILL,   "🟩 Select Store"),
    (BRANCH_FILL,  "🟧 Select Branch"),
    (SUPER_FILL,   "🟪 Select Supervisor"),
    (HOME_FILL,    "🟦 Home Page"),
]:
    ws6.cell(row=row, column=1).fill = fill
    ws6.cell(row=row, column=1).border = BORDER
    ws6.cell(row=row, column=2).value = label
    ws6.cell(row=row, column=2).font = BLACK_NORM
    row += 1

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 18
ws6.column_dimensions["C"].width = 30
ws6.column_dimensions["D"].width = 52
ws6.column_dimensions["E"].width = 16
ws6.column_dimensions["F"].width = 18
ws6.column_dimensions["G"].width = 22


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7: Chat Page — ยืนยัน selector จาก DOM inspection 2026-06-25
# ══════════════════════════════════════════════════════════════════════════════
CHAT_FILL   = PatternFill("solid", fgColor="1A237E")   # deep indigo — chat title
CHAT_ROW_A  = PatternFill("solid", fgColor="E8EAF6")   # light indigo
CHAT_ROW_B  = PatternFill("solid", fgColor="F3E5F5")   # light purple
CHAT_ROW_C  = PatternFill("solid", fgColor="E0F2F1")   # light teal (file/attach)
CHAT_ROW_D  = PatternFill("solid", fgColor="FFF3E0")   # light orange (bubble)
CHAT_ROW_E  = PatternFill("solid", fgColor="FCE4EC")   # light red (close/danger)
CHAT_ROW_F  = PatternFill("solid", fgColor="F1F8E9")   # light green (PASS)
UNVERIFIED  = PatternFill("solid", fgColor="FFF9C4")   # yellow — not yet verified

ws7 = wb.create_sheet("Chat Page")

style_title(ws7, 1, 7, CHAT_FILL,
    "Telepharmacy CMS — Chat Page Selectors  (ยืนยันจาก DOM inspection 2026-06-25)", 12)
ws7.merge_cells("A2:G2")
c = ws7["A2"]
c.value = ("Chat Zone: /home (split-panel) | Zone 1=Queue List | Zone 2=Chat | "
           "Zone 3=Patient Info | Zone 4=Prescription | Zone 5=POS")
c.fill = PatternFill("solid", fgColor="37474F")
c.font = Font(color="ECEFF1", size=9)
c.alignment = CENTER

HEADERS_CHAT = ["หมวด", "Element / ฟีเจอร์", "Selector (Playwright / CSS)",
                "Tag จริง", "Class จริง (ตัดสั้น)", "สถานะ", "หมายเหตุ"]
style_header(ws7, 3, HEADERS_CHAT)

# (section, element, selector, tag, class_snippet, verified, remark)
CHAT_DATA = [

    # ── Layout / Navigation ───────────────────────────────────────────────────
    ("__SECTION__", "LAYOUT  (Grid 5 Columns)", "", "", "", "", ""),

    ("Layout",
     "Root container",
     "div[class*='h-screen'][class*='flex-col']",
     "div", "h-screen flex flex-col bg-slate-100 relative",
     "✅ ยืนยัน", "wrapper ทั้งหน้า"),

    ("Layout",
     "Grid 5 cols",
     "div[class*='grid-cols']",
     "div", "h-full grid grid-cols-[300px_1fr_1fr_2fr] gap-0",
     "✅ ยืนยัน", "Zone1=300px, Zone2=1fr, Zone3=1fr, Zone4=2fr"),

    ("Layout",
     "Zone 1 – Queue List panel",
     "div[class*='bg-white'][class*='flex-col'][class*='overflow-hidden']:first-child",
     "div", "border-r border-slate-200 bg-white flex flex-col overflow-hidden",
     "✅ ยืนยัน", "คอลัมน์ซ้ายสุด — รายการคิวผู้ป่วย"),

    ("Layout",
     "Zone 2 – Chat panel",
     "div[class*='bg-slate-50'][class*='flex-col'][class*='overflow-hidden']",
     "div", "border-r border-slate-200 bg-slate-50 flex flex-col overflow-hidden",
     "✅ ยืนยัน", "คอลัมน์ chat messages"),

    # ── Queue List (Zone 1) ───────────────────────────────────────────────────
    ("__SECTION__", "ZONE 1 — QUEUE LIST", "", "", "", "", ""),

    ("Zone 1",
     "Patient card (clickable row)",
     "div.flex.items-start.gap-3",
     "div", "flex items-start gap-3",
     "✅ ยืนยัน", "คลิกเพื่อเปิด chat ใน Zone 2 — URL ยังอยู่ที่ /home"),

    ("Zone 1",
     "Queue scrollable list",
     "div[class*='scrollbar-thin'][class*='space-y-2']",
     "div", "scrollbar-thin flex-1 space-y-2 overflow-y-auto bg-slate-50/40 p-3",
     "✅ ยืนยัน", "container ของ patient cards"),

    ("Zone 1",
     "Status badge — WAITING",
     "span:has-text('WAITING')",
     "span", "(dynamic — badge text)",
     "✅ ยืนยัน", ""),

    ("Zone 1",
     "Status badge — ACTIVE",
     "span:has-text('ACTIVE')",
     "span", "(dynamic — badge text)",
     "✅ ยืนยัน", ""),

    ("Zone 1",
     "Status badge — CLOSED",
     "span:has-text('CLOSED')",
     "span", "(dynamic — badge text)",
     "✅ ยืนยัน", "ปรากฏหลัง Close Encounter"),

    ("Zone 1",
     "Status badge — PAUSED",
     "span:has-text('PAUSED')",
     "span", "(dynamic — badge text)",
     "✅ ยืนยัน", "ปรากฏหลัง Pause"),

    # ── Chat Header (Zone 2 top) ───────────────────────────────────────────────
    ("__SECTION__", "ZONE 2 — CHAT HEADER", "", "", "", "", ""),

    ("Chat Header",
     "Encounter ID ('Encounter #47')",
     "div:has-text('Encounter #')",
     "div", "(container ที่มีข้อความ 'Encounter #XX')",
     "✅ ยืนยัน", "ข้อความจริงใน header: 'Encounter #47'"),

    ("Chat Header",
     "Encounter timestamp (14:42)",
     "div[class*='text-slate-400'][class*='text-[11px]']",
     "div", "text-[11px] text-slate-400 (approximate)",
     "⚠️ ตรวจเพิ่ม", "อาจตรวจด้วย body text regex: /\\d{1,2}:\\d{2}\\s*น\\./"),

    ("Chat Header",
     "Pause button",
     "button:has-text('Pause')",
     "button", "flex h-9 flex-1 items-center justify-center gap-1.5 rounded-lg border border-sla…",
     "✅ ยืนยัน", "ปุ่มซ้าย ใน bottom bar ของ Zone 2"),

    ("Chat Header",
     "Close Encounter button",
     "button:has-text('Close Encounter')",
     "button", "flex h-9 flex-1 items-center justify-center gap-1.5 rounded-lg text-xs … (สีแดง)",
     "✅ ยืนยัน", "ปุ่มขวา — สีแดง"),

    # ── Chat Message Area (Zone 2 middle) ─────────────────────────────────────
    ("__SECTION__", "ZONE 2 — MESSAGE AREA", "", "", "", "", ""),

    ("Message Area",
     "Scrollable message container",
     "div[class*='scrollbar-thin'][class*='space-y-3']",
     "div", "scrollbar-thin flex-1 space-y-3 overflow-y-auto bg-slate-50 p-4",
     "✅ ยืนยัน", "container ที่บรรจุ message bubbles"),

    ("Message Area",
     "Date badge divider",
     "div[class*='rounded-full'][class*='ring-1']",
     "div", "rounded-full bg-white px-3 py-1 text-[11px] font-medium text-slate-500 ring-1 ring-slate-200",
     "✅ ยืนยัน", "แสดงวันที่กั้นกลาง เช่น '25 มิ.ย. 2569'"),

    ("Message Area",
     "Outbound bubble (operator)",
     "div[class*='justify-end'] div[class*='bg-brand-600']",
     "div", "rounded-2xl px-4 py-2.5 shadow-sm bg-brand-600 text-white rounded-br-md",
     "✅ ยืนยัน", "bubble ฝั่งขวา — พื้นหลังสีน้ำเงิน (brand-600)"),

    ("Message Area",
     "Outbound container (flex-end wrapper)",
     "div[class*='justify-end']",
     "div", "flex justify-end",
     "✅ ยืนยัน", ""),

    ("Message Area",
     "Message text content",
     "div[class*='whitespace-pre-wrap']",
     "div", "whitespace-pre-wrap text-sm leading-relaxed",
     "✅ ยืนยัน", ""),

    ("Message Area",
     "Message timestamp (11:35 น.)",
     "div[class*='text-[11px]'][class*='text-slate-400']",
     "div", "mt-1 px-1 text-[11px] text-slate-400 text-right",
     "✅ ยืนยัน", ""),

    ("Message Area",
     "Inbound bubble (patient)",
     "div[class*='justify-start'] div[class*='rounded-2xl']",
     "div", "(opposite of outbound — justify-start)",
     "⚠️ อนุมาน", "ยังไม่มี inbound message ในขณะ test — ต้องส่งจาก LIFF"),

    # ── Chat Input Area (Zone 2 bottom) ───────────────────────────────────────
    ("__SECTION__", "ZONE 2 — INPUT AREA", "", "", "", "", ""),

    ("Input Area",
     "Chat message input",
     "input[placeholder*='พิมพ์ข้อความ']",
     "input", "h-10 flex-1 rounded-lg border border-slate-200 bg-slate-50 px-3.5 text-sm",
     "✅ ยืนยัน", "⚠️ เป็น input[type=text] ไม่ใช่ textarea!"),

    ("Input Area",
     "Send button (icon, square)",
     "button[class*='bg-brand-600'][class*='h-10'][class*='w-10']",
     "button",
     "flex h-10 w-10 shrink-0 items-center justify-center rounded-lg bg-brand-600 … disabled:opacity-40",
     "✅ ยืนยัน",
     "disabled=true เมื่อ input ว่าง | enabled เมื่อมีข้อความ"),

    ("Input Area",
     "Attach file (label, not button!)",
     "label[class*='shrink-0'][class*='cursor-pointer']",
     "label", "shrink-0 cursor-pointer",
     "✅ ยืนยัน", "⚠️ เป็น <label> ไม่ใช่ <button>! — triggers file input"),

    ("Input Area",
     "File input (hidden)",
     "input[type='file']",
     "input", "hidden",
     "✅ ยืนยัน", "display:none — trigger ผ่าน label"),

    ("Input Area",
     "Mic button",
     "(ไม่มีใน UI ปัจจุบัน)",
     "-", "-",
     "❌ ไม่มี", "ไม่พบ mic button ใน DOM — send disabled/enabled แทน"),

    # ── Close Encounter Dialog ─────────────────────────────────────────────────
    ("__SECTION__", "CONFIRM DIALOG (Close Encounter)", "", "", "", "", ""),

    ("Dialog",
     "Dialog container",
     "[role='dialog']",
     "div",
     "relative flex max-h-[90vh] w-full flex-col rounded-2xl bg-white shadow-2xl ring-1 … animate-fade-in max-w-md",
     "✅ ยืนยัน", "role='dialog' aria-modal='true'"),

    ("Dialog",
     "Dialog title",
     "[role='dialog'] h3",
     "h3", "text-lg font-bold tracking-tight text-slate-900",
     "✅ ยืนยัน", "ข้อความ: 'ยืนยันการจบ Session'"),

    ("Dialog",
     "Cancel button",
     "[role='dialog'] button:has-text('ยกเลิก')",
     "button", "… bg-white hover:bg-slate-50 … text-slate-700 border-slate-200 … h-9 px-4 text-sm …",
     "✅ ยืนยัน", ""),

    ("Dialog",
     "Confirm button (ยืนยันจบ Session)",
     "[role='dialog'] button:has-text('ยืนยันจบ Session')",
     "button", "… bg-danger-600 hover:bg-danger-500 … text-white … h-9 px-4 text-sm …",
     "✅ ยืนยัน", "สีแดง bg-danger-600"),

    # ── After Close ───────────────────────────────────────────────────────────
    ("__SECTION__", "AFTER CLOSE ENCOUNTER", "", "", "", "", ""),

    ("Post-Close",
     "Queue card CLOSED badge",
     "span:has-text('CLOSED')",
     "span", "(status badge in Zone 1)",
     "✅ ยืนยัน", "ปรากฏใน Zone 1 หลังปิด encounter"),

    ("Post-Close",
     "Chat input blocked",
     "input[placeholder*='พิมพ์ข้อความ'][disabled]",
     "input", "(disabled attribute)",
     "⚠️ ตรวจเพิ่ม", "อาจ disabled หรือ hidden ขึ้นกับ implementation"),
]

# Write sheet
row = 4
SECTION_FILLS_CHAT = {
    "LAYOUT": PatternFill("solid", fgColor="1A237E"),
    "ZONE 1": PatternFill("solid", fgColor="2E7D32"),
    "ZONE 2": PatternFill("solid", fgColor="1565C0"),
    "CONFIRM": PatternFill("solid", fgColor="B71C1C"),
    "AFTER": PatternFill("solid", fgColor="6A1B9A"),
}

for record in CHAT_DATA:
    if record[0] == "__SECTION__":
        section_text = record[1]
        sec_fill = next(
            (f for k, f in SECTION_FILLS_CHAT.items() if k.upper() in section_text.upper()),
            PatternFill("solid", fgColor="37474F")
        )
        add_section_label(ws7, row, 7, f"  {section_text}", sec_fill)
        row += 1
        continue

    _, element, selector, tag, cls, verified, remark = record
    # Section = record[0] used for row fill color
    section = record[0]
    if "Input" in section or "Zone 2" in section:
        row_fill = CHAT_ROW_A
    elif "Zone 1" in section or "Queue" in section:
        row_fill = CHAT_ROW_B
    elif "Dialog" in section:
        row_fill = CHAT_ROW_E
    elif "Post" in section or "After" in section:
        row_fill = CHAT_ROW_F
    elif "Layout" in section:
        row_fill = EVEN_FILL
    else:
        row_fill = ODD_FILL

    if "⚠️" in verified or "อนุมาน" in verified:
        row_fill = UNVERIFIED
    elif "❌" in verified:
        row_fill = PatternFill("solid", fgColor="FFCCBC")

    values = [section, element, selector, tag, cls, verified, remark]
    for col, val in enumerate(values, start=1):
        c = ws7.cell(row=row, column=col, value=val)
        c.fill = row_fill
        c.font = BLACK_NORM
        c.alignment = LEFT
        c.border = BORDER
    row += 1

ws7.column_dimensions["A"].width = 14
ws7.column_dimensions["B"].width = 30
ws7.column_dimensions["C"].width = 52
ws7.column_dimensions["D"].width = 8
ws7.column_dimensions["E"].width = 52
ws7.column_dimensions["F"].width = 14
ws7.column_dimensions["G"].width = 40

ws7.freeze_panes = "A4"


# ── Save ─────────────────────────────────────────────────────────────────────
OUTPUT = "telepharmacy-selector-discovery/Telepharmacy_CMS_Selectors.xlsx"
wb.save(OUTPUT)
import sys
out = sys.stdout.buffer if hasattr(sys.stdout, 'buffer') else None
def p(text):
    if out:
        out.write((text + "\n").encode("utf-8"))
    else:
        print(text.encode("utf-8", errors="replace").decode("ascii", errors="replace"))

p(f"SAVED: {OUTPUT}")
p(f"   Sheet 'Login Flow'          : {len([x for x in FLOW_DATA if x[0] != '__SECTION__'])} rows")
p(f"   Sheet 'Login Page'          : {len(LOGIN_DETAIL)} rows")
p(f"   Sheet 'Select Store+Branch' : {len([x for x in STORE_BRANCH_DATA if x[0] != '__SECTION__'])} rows")
p(f"   Sheet 'Supervisor+Home'     : {len([x for x in SUPER_HOME_DATA if x[0] != '__SECTION__'])} rows")
p(f"   Sheet 'Code Snippets'       : {len(SNIPPETS)} snippets")
p(f"   Sheet 'Summary'             : {len(SUMMARY_DATA)} selectors total")
p(f"   Sheet 'Chat Page'           : {len([x for x in CHAT_DATA if x[0] != '__SECTION__'])} rows (verified 2026-06-25)")
