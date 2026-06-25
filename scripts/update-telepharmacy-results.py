"""
update-telepharmacy-results.py
อ่านผลจาก JSON ทุกไฟล์แล้วเขียนกลับเข้า Arincare_Telepharmacy_TestCases.xlsx

col 11 = Actual Result
col 12 = Status  (PASS/FAIL/SKIP)
col 13 = Remark + screenshot list
"""

import json, os, sys
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

BASE  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL = os.path.join(BASE, "Arincare_Telepharmacy_TestCases.xlsx")

JSON_FILES = [
    os.path.join(BASE, "test-results-login-final.json"),
    os.path.join(BASE, "test-results-home.json"),
    os.path.join(BASE, "test-results-queue.json"),
    os.path.join(BASE, "test-results-chat.json"),
    os.path.join(BASE, "test-results-call.json"),
]

PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
SKIP_FILL  = PatternFill("solid", fgColor="FFEB9C")
PASS_FONT  = Font(bold=True, color="276221", size=10)
FAIL_FONT  = Font(bold=True, color="9C0006", size=10)
SKIP_FONT  = Font(bold=True, color="9C6500", size=10)
NORM_FONT  = Font(size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

def p(text):
    sys.stdout.buffer.write((text + "\n").encode("utf-8"))

# ── Load results ────────────────────────────────────────────────────────────────
results = {}
for fpath in JSON_FILES:
    if os.path.exists(fpath):
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        for r in data:
            results[r["id"]] = r
        p(f"Loaded: {os.path.basename(fpath)}  ({len(data)} records)")
    else:
        p(f"SKIP : {os.path.basename(fpath)} — ไม่พบไฟล์")

if not results:
    p("ERROR: ไม่พบ result JSON — รันเทสก่อน")
    sys.exit(1)

p(f"\nTotal: {len(results)} results loaded")

# ── Open workbook ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["Test Cases"]

# ── Build TC-ID → row map ───────────────────────────────────────────────────────
row_map = {}
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, 1).value
    if tc_id:
        row_map[str(tc_id).strip()] = row

p(f"Excel: {len(row_map)} rows in 'Test Cases'")

# ── Write results ───────────────────────────────────────────────────────────────
updated  = 0
skipped  = []
counters = {"PASS": 0, "FAIL": 0, "SKIP": 0}

for tc_id, r in sorted(results.items()):
    row = row_map.get(tc_id)
    if row is None:
        skipped.append(tc_id)
        continue

    status = r.get("status", "")
    actual = r.get("actualResult", "")
    remark = r.get("remark", "")
    shots  = r.get("screenshots", [])

    c11 = ws.cell(row, 11)
    c11.value     = actual
    c11.font      = NORM_FONT
    c11.alignment = WRAP_ALIGN

    c12 = ws.cell(row, 12)
    c12.value = status
    if status == "PASS":
        c12.fill = PASS_FILL; c12.font = PASS_FONT
    elif status == "FAIL":
        c12.fill = FAIL_FILL; c12.font = FAIL_FONT
    else:
        c12.fill = SKIP_FILL; c12.font = SKIP_FONT
    c12.alignment = Alignment(horizontal="center", vertical="center")

    ss_text = ""
    if shots:
        ss_text = "\n\nScreenshots:\n" + "\n".join(f"  • {s}" for s in shots[:5])
        if len(shots) > 5:
            ss_text += f"\n  ... +{len(shots)-5} more"

    c13 = ws.cell(row, 13)
    c13.value     = remark + ss_text
    c13.font      = NORM_FONT
    c13.alignment = WRAP_ALIGN

    ws.row_dimensions[row].height = max(30, 15 + min(len(shots), 5) * 12)

    counters[status] = counters.get(status, 0) + 1
    updated += 1
    icon = "✅" if status == "PASS" else "❌" if status == "FAIL" else "⏭️"
    p(f"  {icon} row {row:3d} | {tc_id} | {status}")

# ── Update สรุปภาพรวม sheet ─────────────────────────────────────────────────────
if "สรุปภาพรวม" in wb.sheetnames:
    ws_sum = wb["สรุปภาพรวม"]
    run_dt = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Write summary block starting at A1
    summary_rows = [
        ("รายงานผลการทดสอบ Arincare Telepharmacy CMS", ""),
        ("วันที่อัพเดท", run_dt),
        ("", ""),
        ("สถานะ", "จำนวน"),
        ("PASS",  counters.get("PASS", 0)),
        ("FAIL",  counters.get("FAIL", 0)),
        ("SKIP",  counters.get("SKIP", 0)),
        ("รวม",   updated),
        ("", ""),
        ("โมดูลที่รัน", ""),
    ]
    # Add per-module breakdown
    modules: dict[str, dict] = {}
    for tc_id, r in results.items():
        mod = tc_id.split("-")[1] if "-" in tc_id else "?"
        if mod not in modules:
            modules[mod] = {"PASS": 0, "FAIL": 0, "SKIP": 0}
        st = r.get("status", "SKIP")
        modules[mod][st] = modules[mod].get(st, 0) + 1

    for mod, mc in sorted(modules.items()):
        summary_rows.append((
            f"TC-{mod}-*",
            f"PASS={mc.get('PASS',0)}  FAIL={mc.get('FAIL',0)}  SKIP={mc.get('SKIP',0)}"
        ))

    hdr_font  = Font(bold=True, size=11)
    stat_fill = {
        "PASS": PatternFill("solid", fgColor="C6EFCE"),
        "FAIL": PatternFill("solid", fgColor="FFC7CE"),
        "SKIP": PatternFill("solid", fgColor="FFEB9C"),
    }

    for i, (col_a, col_b) in enumerate(summary_rows, start=1):
        ca = ws_sum.cell(i, 1); ca.value = col_a
        cb = ws_sum.cell(i, 2); cb.value = col_b
        if col_a in ("PASS", "FAIL", "SKIP"):
            ca.fill = stat_fill[col_a]
            cb.fill = stat_fill[col_a]
            ca.font = Font(bold=True, size=10)
        elif i == 1:
            ca.font = Font(bold=True, size=13)
        elif col_a in ("สถานะ", "โมดูลที่รัน"):
            ca.font = hdr_font; cb.font = hdr_font

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 40
    p(f"\nอัพเดท สรุปภาพรวม sheet แล้ว")

# ── Save ────────────────────────────────────────────────────────────────────────
wb.save(EXCEL)

p(f"\n{'='*50}")
p(f"Updated : {updated} rows")
p(f"PASS={counters.get('PASS',0)}  FAIL={counters.get('FAIL',0)}  SKIP={counters.get('SKIP',0)}")
p(f"Skipped : {len(skipped)} (ไม่มีใน Excel: {', '.join(skipped) if skipped else 'none'})")
p(f"Saved   : {EXCEL}")
